"""
Router für Import-Batches.

Endpunkte:
  GET  /api/imports/              — alle Batches (optional gefiltert)
  POST /api/imports/              — neuen Import starten
  GET  /api/imports/{id}          — Batch-Details mit Dokumentliste
  GET  /api/imports/{id}/export   — Excel-Export aller Dokumente des Batches
"""

import asyncio
import io
import logging
from datetime import date
from pathlib import Path
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app import crud
from app.database import SessionLocal, get_db
from app.schemas.import_batch import (
    ImportBatchAutomationUpdate,
    ImportBatchCreate,
    ImportBatchRead,
    ImportBatchWithDocuments,
)
from app.services.import_service import (
    _run_import_io,
    list_pdf_files,
    run_import,
    validate_import_path,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/imports", tags=["Imports"])


# ─── Sync DB-Hilfsfunktionen (laufen via asyncio.to_thread) ─────────────────

def _db_get_source_filenames(batch_id: int) -> list[str]:
    """
    Gibt die original_filename aller erfolgreich kopierten Dokumente zurück.
    Direkte Query statt Relationship-Zugriff für maximale Zuverlässigkeit.
    """
    from app.models.document import Document as _Doc
    db = SessionLocal()
    try:
        rows = (
            db.query(_Doc.original_filename)
            .filter(
                _Doc.batch_id == batch_id,
                _Doc.stored_filename.isnot(None),
                _Doc.original_filename.isnot(None),
            )
            .all()
        )
        names = [r.original_filename for r in rows]
        logger.info(
            "Batch #%d: %d Quelldatei(en) in DB gefunden", batch_id, len(names)
        )
        return names
    finally:
        db.close()


def _db_get_analyze_setup(
    batch_id: int,
    ai_config_id: int | None,
) -> dict | None:
    """
    Löst KI-Config und Dokument-IDs für die KI-Analyse auf.
    Gibt None zurück, wenn keine KI-Konfiguration gefunden wurde.

    Löst absichtlich NICHT den Systemprompt-Inhalt auf — der Prompt wird nur als
    system_prompt_id (nicht als aufgelöster Text) in die Task-Payload geschrieben
    und erst von _db_analyze_read() im Worker frisch aus der DB gelesen. So greift
    eine nachträgliche Bearbeitung des Prompts auch bei bereits wartenden Tasks.
    """
    db = SessionLocal()
    try:
        if ai_config_id:
            resolved_config = crud.ai_config.get_by_id(db, ai_config_id)
        else:
            resolved_config = crud.ai_config.get_default(db)

        if resolved_config is None:
            logger.warning(
                "Batch #%d: Keine KI-Konfiguration gefunden — KI-Analyse übersprungen",
                batch_id,
            )
            return None

        batch = crud.import_batch.get_by_id(db, batch_id)
        doc_ids = [d.id for d in (batch.documents if batch else []) if d.stored_filename]

        return {
            "ai_config_id": resolved_config.id,
            "doc_ids": doc_ids,
        }
    finally:
        db.close()


def _db_enqueue_for_analysis(doc_ids: list[int], system_prompt_id: int | None) -> int:
    """
    Reiht Dokumente zur KI-Analyse in die Worker-Queue ein.
    Verhindert Duplikate: Dokumente mit bereits laufendem/wartendem Task werden übersprungen.
    Gibt die Anzahl neu eingereihter Dokumente zurück.

    system_prompt_id wird unverändert in die Task-Payload geschrieben (keine
    Text-Auflösung hier) — siehe _db_get_analyze_setup().
    """
    import uuid
    from sqlalchemy import text as _text
    from app.models.workflow_task import WorkflowTask
    db = SessionLocal()
    count = 0
    try:
        for doc_id in doc_ids:
            existing = db.execute(
                _text(
                    "SELECT id FROM workflow_tasks "
                    "WHERE payload->>'document_id' = :doc_id "
                    "AND status IN ('pending', 'in_progress') LIMIT 1"
                ),
                {"doc_id": str(doc_id)},
            ).first()
            if existing:
                logger.debug("Dokument #%d bereits in Worker-Queue — übersprungen", doc_id)
                continue
            task = WorkflowTask(
                workflow_id=str(uuid.uuid4()),
                payload={
                    "kind": "process_document",
                    "document_id": doc_id,
                    "system_prompt_id": system_prompt_id,
                },
                status="pending",
            )
            db.add(task)
            count += 1
        db.commit()
        return count
    except Exception as exc:
        logger.error("Fehler beim Einreihen der Analyse-Tasks: %s", exc)
        try:
            db.rollback()
        except Exception:
            pass
        return 0
    finally:
        db.close()


def _sync_delete_source_files(import_folder: str, original_names: list[str]) -> tuple[int, int]:
    """
    Sync: Löscht Quelldateien aus dem Import-Ordner.
    Gibt (deleted, failed) zurück.

    Ist der Import-Unterordner danach leer, wird er ebenfalls entfernt (analog zum
    Verhalten beim Batch-Löschen in delete_import()) — außer es handelt sich um
    IMPORT_BASE_PATH selbst (kein Unterordner angegeben), das bleibt immer erhalten.
    """
    from app.config import settings as _settings

    folder = Path(import_folder)
    deleted, failed = 0, 0
    logger.info(
        "Lösche Quelldateien aus '%s': %d Datei(en)", import_folder, len(original_names)
    )
    for name in original_names:
        src = folder / name
        try:
            src.unlink()  # Direkt löschen — FileNotFoundError bei fehlendem File
            deleted += 1
            logger.info("Quelldatei gelöscht: %s", src)
        except FileNotFoundError:
            logger.warning("Quelldatei nicht gefunden (bereits gelöscht?): %s", src)
        except Exception as exc:
            failed += 1
            logger.error("Konnte Quelldatei nicht löschen %s: %s", src, exc)

    # Leeren Import-Unterordner entfernen (nie IMPORT_BASE_PATH selbst löschen)
    try:
        base = Path(_settings.import_base_path).resolve()
        resolved = folder.resolve()
        if resolved != base and resolved.exists() and not any(resolved.iterdir()):
            resolved.rmdir()
            logger.info("Leerer Import-Ordner gelöscht: %s", resolved)
    except Exception as exc:
        logger.warning("Konnte leeren Import-Ordner nicht löschen %s: %s", folder, exc)

    return deleted, failed


async def _delete_source_files(batch_id: int, import_folder: str) -> None:
    """
    Löscht die Original-PDFs aus dem Import-Ordner, die erfolgreich kopiert wurden.
    Es werden nur Dateien gelöscht, für die ein DB-Eintrag mit stored_filename existiert.
    """
    # DB-Abfrage im Thread
    original_names = await asyncio.to_thread(_db_get_source_filenames, batch_id)

    if not original_names:
        logger.warning(
            "Batch #%d: Keine Quelldateien in DB gefunden — nichts zu löschen", batch_id
        )
        return

    logger.info(
        "Batch #%d: %d Quelldatei(en) werden aus '%s' gelöscht",
        batch_id, len(original_names), import_folder,
    )

    # Filesystem-Operationen im Thread (einfaches asyncio.to_thread, kein spezieller Pool nötig)
    deleted, failed = await asyncio.to_thread(
        _sync_delete_source_files, import_folder, original_names
    )

    logger.info(
        "Batch #%d: Quelldateien gelöscht=%d, fehlgeschlagen=%d", batch_id, deleted, failed
    )


async def _import_and_delete(batch_id: int, import_folder: str) -> None:
    """Import durchführen und danach Quelldateien löschen (ohne KI-Analyse)."""
    try:
        await run_import(batch_id)
        await _delete_source_files(batch_id, import_folder)
    except Exception as exc:
        logger.error(
            "Batch #%d: Fehler in _import_and_delete: %s", batch_id, exc, exc_info=True
        )


async def _import_then_analyze(
    batch_id: int,
    import_folder: str,
    ai_config_id: int | None = None,
    system_prompt_id: int | None = None,
    delete_source_files: bool = False,
) -> None:
    """
    Führt zuerst den Import durch, reiht danach alle importierten Dokumente
    zur KI-Analyse in die Worker-Queue ein.

    Alle DB- und Filesystem-Operationen laufen in Thread-Pools (nicht-blockierend).
    Die Analyse wird vom WorkerPool übernommen — keine direkte _run_analysis-Ausführung.
    Das verhindert Race-Conditions zwischen Import-Pfad und Worker-Queue.
    """
    try:
        # 1. Import abwarten
        await run_import(batch_id)

        # 2. Quelldateien löschen (falls gewünscht), bevor KI läuft
        if delete_source_files:
            await _delete_source_files(batch_id, import_folder)

        # 3. Dokument-IDs für diesen Batch ermitteln (im Thread)
        setup = await asyncio.to_thread(_db_get_analyze_setup, batch_id, ai_config_id)
        if setup is None:
            return

        doc_ids: list[int] = setup["doc_ids"]
        if not doc_ids:
            logger.info("Batch #%d: Keine Dokumente für KI-Analyse", batch_id)
            return

        # 4. Dokumente in Worker-Queue einreihen (statt direkter Analyse)
        #    Verhindert Duplikate falls Dokumente bereits in der Queue sind.
        enqueued = await asyncio.to_thread(_db_enqueue_for_analysis, doc_ids, system_prompt_id)
        logger.info(
            "Batch #%d: %d/%d Dokument(e) zur KI-Analyse in Worker-Queue eingereiht",
            batch_id, enqueued, len(doc_ids),
        )

    except Exception as exc:
        logger.error(
            "Batch #%d: Fehler in _import_then_analyze: %s", batch_id, exc, exc_info=True
        )


@router.get("", response_model=list[ImportBatchRead])
def list_imports(
    company_name: str | None = Query(None, description="Firmenname (Teilstring-Suche)"),
    year: int | None = Query(None, description="Importjahr (exakt)"),
    db: Session = Depends(get_db),
):
    """
    Gibt alle Import-Batches zurück.
    Optional nach Firmenname (Teilstring) und/oder Jahr filtern.
    """
    return crud.import_batch.get_all(db, company_name=company_name, year=year)


@router.post("", response_model=ImportBatchRead, status_code=status.HTTP_201_CREATED)
async def start_import(payload: ImportBatchCreate, db: Session = Depends(get_db)):
    """
    Startet einen neuen Import-Vorgang.

    Der Ordnerpfad wird aus Firma + Jahr konstruiert: IMPORT_BASE_PATH/Firma_Jahr
    Falls der Ordner noch nicht existiert, wird er angelegt.
    Keine KI-Extraktion beim Import – nur Kopieren und Metadaten erfassen.
    """
    # ── Firma + Jahr bestimmen ─────────────────────────────────────────────
    if not payload.company_name or not payload.year:
        raise HTTPException(status_code=400, detail="Firmenname und Jahr sind erforderlich.")
    company_name = payload.company_name
    year = payload.year

    # ── Import-Pfad: IMPORT_BASE_PATH + optionaler Unterordner ──────────────
    from app.config import settings as _settings
    base = _settings.import_base_path
    subfolder = (payload.subfolder or "").strip().strip("/\\")
    folder_path = f"{base}/{subfolder}" if subfolder else base
    payload = payload.model_copy(update={"folder_path": folder_path})

    import functools
    try:
        validated_path = await asyncio.to_thread(validate_import_path, folder_path)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    # ── PDF-Prüfung ────────────────────────────────────────────────────────
    pdf_files = await asyncio.to_thread(list_pdf_files, validated_path)
    if not pdf_files:
        raise HTTPException(
            status_code=400,
            detail=f"Keine PDF-Dateien im Import-Ordner gefunden.",
        )

    # ── Batch in DB anlegen ────────────────────────────────────────────────
    batch = crud.import_batch.create(
        db=db,
        data=payload,
        company_name=company_name,
        year=year,
    )
    logger.info("Import-Batch #%d erstellt: %s_%d (%d PDFs)", batch.id, company_name, year, len(pdf_files))

    if payload.analyze_after_import:
        asyncio.create_task(_import_then_analyze(
            batch_id=batch.id,
            import_folder=folder_path,
            ai_config_id=payload.ai_config_id,
            system_prompt_id=payload.system_prompt_id,
            delete_source_files=payload.delete_source_files,
        ))
    elif payload.delete_source_files:
        asyncio.create_task(_import_and_delete(batch.id, folder_path))
    else:
        asyncio.create_task(run_import(batch.id))
    return batch


@router.get("/{batch_id}/status", response_model=ImportBatchRead)
def get_import_status(batch_id: int, db: Session = Depends(get_db)):
    """
    Gibt nur den Status eines Import-Batches zurück — OHNE Dokumentliste.
    Für leichtgewichtiges Polling während eines laufenden Imports.
    Enthält total_documents für die Fortschrittsanzeige.
    """
    from app.models.import_batch import ImportBatch as ImportBatchModel
    from app.models.document import Document
    from sqlalchemy import func, select as sa_select

    obj = db.get(ImportBatchModel, batch_id)
    if obj is None:
        raise HTTPException(status_code=404, detail="Import-Batch nicht gefunden")

    total_docs: int = db.execute(
        sa_select(func.count()).where(Document.batch_id == batch_id)
    ).scalar_one()

    result = ImportBatchRead.model_validate(obj)
    result.total_documents = total_docs
    return result


@router.patch("/{batch_id}/automation", response_model=ImportBatchRead)
def update_import_automation(
    batch_id: int,
    payload: ImportBatchAutomationUpdate,
    db: Session = Depends(get_db),
):
    """
    Aktiviert/deaktiviert Ordner-Sync und automatischen Export nachträglich für einen
    bestehenden Import-Batch (bei der Erstellung gesetzte Werte sind kein Dauerzustand).
    Nur übergebene Felder werden geändert (None = unverändert lassen).
    """
    from app.models.import_batch import ImportBatch as _Batch

    batch = db.get(_Batch, batch_id)
    if batch is None:
        raise HTTPException(status_code=404, detail="Import-Batch nicht gefunden")

    if payload.folder_sync is not None:
        batch.folder_sync = payload.folder_sync
    if payload.auto_export is not None:
        batch.auto_export = payload.auto_export

    db.commit()
    db.refresh(batch)

    result = ImportBatchRead.model_validate(batch)
    result.total_documents = len(batch.documents)
    return result


@router.get("/{batch_id}/ki-stats")
def get_batch_ki_stats(batch_id: int, db: Session = Depends(get_db)):
    """Aggregierte KI-Statistiken für alle Dokumente eines Import-Batches."""
    from sqlalchemy import func as sqlfunc
    from app.models.document_token_count import DocumentTokenCount as _TC
    from app.models.document import Document as _Doc

    row = (
        db.query(
            sqlfunc.sum(_TC.input_token_count + _TC.output_token_count).label("total_tokens"),
            sqlfunc.sum(_TC.time_spent_seconds).label("total_duration_seconds"),
        )
        .join(_Doc, _TC.document_id == _Doc.id)
        .filter(_Doc.batch_id == batch_id)
        .first()
    )

    return {
        "total_tokens": int(row.total_tokens or 0),
        "total_duration_seconds": float(row.total_duration_seconds or 0.0),
    }


@router.get("/{batch_id}", response_model=ImportBatchWithDocuments)
def get_import(batch_id: int, db: Session = Depends(get_db)):
    """Gibt einen Import-Batch mit vollständiger Dokumentliste zurück."""
    batch = crud.import_batch.get_by_id(db, batch_id)
    if batch is None:
        raise HTTPException(status_code=404, detail="Import-Batch nicht gefunden")
    return batch


@router.delete("/{batch_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_import(batch_id: int, db: Session = Depends(get_db)):
    """Löscht einen Import-Batch, alle zugehörigen Dokumente und die PDF-Dateien vom Filesystem."""
    import shutil
    from pathlib import Path

    from app.config import settings as _settings

    # Batch mit Dokumenten laden
    batch = crud.import_batch.get_by_id(db, batch_id)
    if batch is None:
        raise HTTPException(status_code=404, detail="Import-Batch nicht gefunden")

    # PDF-Dateien vom Filesystem löschen
    deleted_files = 0
    failed_files = 0
    deleted_folders: set[Path] = set()

    for doc in batch.documents:
        if doc.stored_filename:
            pdf_path = Path(batch.storage_folder_path) / doc.stored_filename
            if pdf_path.exists():
                try:
                    pdf_path.unlink()
                    deleted_files += 1
                    deleted_folders.add(pdf_path.parent)
                    logger.info("Datei gelöscht: %s", pdf_path)
                except Exception as exc:
                    failed_files += 1
                    logger.warning("Fehler beim Löschen von %s: %s", pdf_path, exc)

    # Leere Unterordner ebenfalls löschen
    for folder in deleted_folders:
        try:
            if folder.exists() and not any(folder.iterdir()):
                folder.rmdir()
                logger.info("Leerer Ordner gelöscht: %s", folder)
        except Exception as exc:
            logger.warning("Fehler beim Löschen des Ordners %s: %s", folder, exc)

    if deleted_files or failed_files:
        logger.info(
            "Batch #%d: %d Datei(en) gelöscht, %d Fehler",
            batch_id, deleted_files, failed_files,
        )

    # DB-Eintrag löschen (CASCADE zu Dokumenten, Extraktionen, Positionen)
    crud.import_batch.delete(db, batch_id)


# ─── GET /api/imports/{batch_id}/export ──────────────────────────────────────

def _get_export_config_fields(db: Session) -> tuple[list[str], list[str]]:
    """Liest die aktive Spaltenkonfiguration (ExportConfig, Singleton id=1)."""
    from app.models.export_config import (
        ExportConfig as _ExportConfig,
        INVOICE_FIELDS_DEFAULT,
        POSITION_FIELDS_DEFAULT,
    )

    cfg = db.get(_ExportConfig, 1)
    invoice_fields  = list(cfg.invoice_fields)  if cfg and cfg.invoice_fields  else list(INVOICE_FIELDS_DEFAULT)
    position_fields = list(cfg.position_fields) if cfg and cfg.position_fields else list(POSITION_FIELDS_DEFAULT)
    return invoice_fields, position_fields


def _query_batch_documents(db: Session, batch_id: int, since=None) -> list:
    """
    Lädt alle nicht-gelöschten Dokumente eines Batches (inkl. Extraktion + Positionen).

    since: falls gesetzt, werden nur abgeschlossene Dokumente zurückgegeben, die NACH
    diesem Zeitpunkt zuletzt geändert wurden (document.updated_at > since) — Basis für
    den inkrementellen Export ("/export/new" und den automatischen Wochen-Export).
    """
    from sqlalchemy.orm import joinedload as _jl
    from app.models.document import Document as _Doc

    query = (
        db.query(_Doc)
        .options(_jl(_Doc.extraction), _jl(_Doc.order_positions))
        .filter(_Doc.batch_id == batch_id, _Doc.soft_deleted == False)  # noqa: E712
    )
    if since is not None:
        query = query.filter(_Doc.status == "done", _Doc.updated_at > since)
    return query.order_by(_Doc.id).all()


@router.get("/{batch_id}/export")
def export_batch_excel(
    batch_id: int,
    date_from: date | None = Query(None, description="Nur Dokumente mit Datum >= date_from (ISO YYYY-MM-DD)"),
    date_to: date | None = Query(None, description="Nur Dokumente mit Datum <= date_to (ISO YYYY-MM-DD)"),
    date_field: Literal["invoice_date", "import_date"] = Query(
        "invoice_date",
        description="Welches Datum für date_from/date_to verwendet wird: "
                     "'invoice_date' (Rechnungsdatum) oder 'import_date' (Importdatum)",
    ),
    db: Session = Depends(get_db),
):
    """
    Exportiert alle nicht-gelöschten Dokumente eines Import-Batches als Excel-Datei.

    Sheet 1 „Rechnungen": Rechnungsdaten pro Dokument (Lieferant, Beträge, Daten …)
    Sheet 2 „Positionen": Alle Bestellpositionen aller Dokumente des Batches

    Welche Spalten erscheinen, wird durch die ExportConfig (Singleton id=1) gesteuert.
    Fehlt der Eintrag, werden alle Felder ausgegeben.

    Optionale Filter date_from/date_to schränken die Dokumente ein. date_field legt fest,
    welches Datum geprüft wird:
      - "invoice_date" (Standard): Rechnungsdatum (InvoiceExtraction.invoice_date).
        Dokumente ohne Rechnungsdatum (z.B. Nicht-Eingangsrechnungen) werden dann
        ausgeschlossen, da ihr Datum nicht bestimmbar ist.
      - "import_date": Importdatum (Document.created_at, in lokaler Zeitzone).
    """
    from zoneinfo import ZoneInfo

    from app.config import settings as _settings
    from app.models.import_batch import ImportBatch as _Batch

    batch = db.get(_Batch, batch_id)
    if batch is None:
        raise HTTPException(status_code=404, detail="Import-Batch nicht gefunden")

    docs = _query_batch_documents(db, batch_id)

    if date_from or date_to:
        local_tz = ZoneInfo(_settings.timezone)

        def _doc_date(doc) -> date | None:
            if date_field == "import_date":
                return doc.created_at.astimezone(local_tz).date() if doc.created_at else None
            return doc.extraction.invoice_date if doc.extraction else None

        def _in_date_range(doc) -> bool:
            d = _doc_date(doc)
            if d is None:
                return False
            if date_from and d < date_from:
                return False
            if date_to and d > date_to:
                return False
            return True
        docs = [d for d in docs if _in_date_range(d)]

    invoice_fields, position_fields = _get_export_config_fields(db)
    excel_bytes = _build_export_excel(batch, docs, invoice_fields, position_fields)
    safe = f"{batch.company_name}_{batch.year}".replace(" ", "_")
    if date_from or date_to:
        safe += f"_{date_from or 'anfang'}_bis_{date_to or 'ende'}"
    filename = f"Export_{safe}.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{batch_id}/export/new")
def export_batch_excel_incremental(batch_id: int, db: Session = Depends(get_db)):
    """
    Exportiert nur Dokumente, die seit dem letzten Abruf fertig analysiert wurden.

    Teilt sich last_exported_at mit dem automatischen Wochen-Export
    (app/worker/export_schedule.py) — ein gemeinsamer Zähler, egal ob manuell oder
    automatisch ausgelöst. Noch nie exportiert (last_exported_at ist None) → alle
    abgeschlossenen Dokumente gelten als neu. Aktualisiert last_exported_at danach,
    auch wenn keine neuen Dokumente gefunden wurden (leere Excel mit nur Kopfzeilen).
    """
    from datetime import datetime, timezone

    from app.models.import_batch import ImportBatch as _Batch

    batch = db.get(_Batch, batch_id)
    if batch is None:
        raise HTTPException(status_code=404, detail="Import-Batch nicht gefunden")

    docs = _query_batch_documents(db, batch_id, since=batch.last_exported_at)
    invoice_fields, position_fields = _get_export_config_fields(db)
    excel_bytes = _build_export_excel(batch, docs, invoice_fields, position_fields)

    batch.last_exported_at = datetime.now(timezone.utc)
    db.commit()

    safe = f"{batch.company_name}_{batch.year}".replace(" ", "_")
    filename = f"Export_{safe}_neu.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_export_excel(
    batch,
    docs,
    invoice_fields: list[str],
    position_fields: list[str],
) -> bytes:
    """
    Erstellt eine formatierte Excel-Datei mit zwei Sheets.

    invoice_fields / position_fields: geordnete Liste aktiver Feld-Schlüssel aus
    ExportConfig — nur diese Spalten werden ausgegeben.
    """
    import json as _json
    from zoneinfo import ZoneInfo

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from app.config import settings as _settings

    _local_tz = ZoneInfo(_settings.timezone)

    wb = Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="2D3748")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    fill_a    = PatternFill("solid", fgColor="FFFFFF")
    fill_b    = PatternFill("solid", fgColor="F0F4F8")
    center    = Alignment(horizontal="center", vertical="center")
    fmt_eur   = '#,##0.00 €'
    fmt_pct   = '0.00"%"'
    fmt_date  = 'DD.MM.YYYY'
    fmt_dt    = 'DD.MM.YYYY HH:MM'

    def _header_row(ws, headers: list[str], row: int = 1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
        ws.row_dimensions[row].height = 28

    def _set_widths(ws, widths: list[int]):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _row_fill(row_idx: int) -> PatternFill:
        return fill_a if row_idx % 2 == 0 else fill_b

    def _f(val):
        return float(val) if val is not None else None

    def _dt(val):
        """Wandelt einen UTC-Zeitstempel in lokale Zeit um (für Excel-Anzeige ohne tzinfo)."""
        if val is None:
            return None
        if getattr(val, "tzinfo", None):
            val = val.astimezone(_local_tz)
            return val.replace(tzinfo=None)
        return val

    def _parse_raw(raw_response) -> dict:
        if not raw_response:
            return {}
        try:
            if isinstance(raw_response, dict):
                return raw_response
            result = _json.loads(raw_response)
            return result if isinstance(result, dict) else {}
        except (_json.JSONDecodeError, TypeError):
            return {}

    def _nested(d: dict, *keys) -> dict:
        """Läuft eine Kette verschachtelter dict-Keys sicher ab, z.B.
        _nested(raw, "lieferant", "anschrift") → {} falls ein Zwischenschritt fehlt."""
        for k in keys:
            d = (d or {}).get(k) or {}
        return d

    # ── Feld-Definitionen: (schlüssel, header, value_fn, format, breite) ─────
    # format: None | "eur" | "pct" | "date" | "datetime" | "center"

    # Rechnungen-Sheet
    INVOICE_DEFS = [
        ("beleg_nr",            "Beleg-Nr.",             lambda d, e, sk: d.id,                                        "center",    10),
        ("dateiname",           "Dateiname",              lambda d, e, sk: d.original_filename,                         None,        36),
        ("status",              "Status",                 lambda d, e, sk: d.status,                                    None,        12),
        ("seiten",              "Seiten",                 lambda d, e, sk: d.page_count or None,                        "center",     8),
        ("rechnungsnr",         "Rechnungsnr.",           lambda d, e, sk: e.invoice_number if e else None,             None,        22),
        ("rechnungsdatum",      "Rechnungsdatum",         lambda d, e, sk: e.invoice_date if e else None,               "date",      14),
        ("faelligkeit",         "Fälligkeit",             lambda d, e, sk: e.due_date if e else None,                   "date",      14),
        ("lieferant",           "Lieferant",              lambda d, e, raw: e.vendor_id if e else None,                 None,        30),
        ("strasse",             "Straße",                 lambda d, e, raw: _nested(raw, "lieferant", "anschrift").get("strasse"),  None,        28),
        ("plz",                 "PLZ",                    lambda d, e, raw: _nested(raw, "lieferant", "anschrift").get("plz"),      None,         8),
        ("ort",                 "Ort",                    lambda d, e, raw: _nested(raw, "lieferant", "anschrift").get("ort"),      None,        20),
        ("ust_id",              "USt-IdNr.",              lambda d, e, raw: _nested(raw, "lieferant").get("ust_id_nr"),             None,        18),
        ("steuernr",            "Steuernr.",              lambda d, e, raw: _nested(raw, "lieferant").get("steuernummer"),          None,        16),
        ("hrb_nr",              "HRB-Nr.",                lambda d, e, raw: _nested(raw, "lieferant").get("hrb_nummer"),            None,        14),
        ("kundennr",            "Kundennr.",              lambda d, e, raw: _nested(raw, "rechnungsdaten").get("kundennummer"),     None,        14),
        ("bestellnr",           "Bestellnr.",             lambda d, e, raw: _nested(raw, "rechnungsdaten").get("bestellnummer"),    None,        16),
        ("bank",                "Bank",                   lambda d, e, raw: _nested(raw, "lieferant", "bankverbindung").get("bank_name"), None,   24),
        ("iban",                "IBAN",                   lambda d, e, raw: _nested(raw, "lieferant", "bankverbindung").get("iban"),      None,   26),
        ("bic",                 "BIC",                    lambda d, e, raw: _nested(raw, "lieferant", "bankverbindung").get("bic"),       None,   12),
        ("gesamtbetrag",        "Gesamtbetrag (€)",       lambda d, e, raw: _f(e.total_amount_brutto) if e else None,   "eur",       17),
        ("rabatt",              "Rabatt (€)",             lambda d, e, raw: _f(e.discount_amount) if e else None,       "eur",       13),
        ("skonto_betrag",       "Skonto (€)",             lambda d, e, raw: _f(e.cash_discount_amount) if e else None,  "eur",       13),
        ("skonto_prozent",      "Skonto (%)",             lambda d, e, raw: _f(_nested(raw, "zahlungsinformationen", "skonto").get("prozent")), "pct", 10),
        ("skonto_frist",        "Skonto Frist (Tage)",    lambda d, e, raw: (
            lambda sk: (
                int(sk["frist_tage"]) if isinstance(sk.get("frist_tage"), float) and sk["frist_tage"] == int(sk["frist_tage"])
                else sk.get("frist_tage")
            )
        )(_nested(raw, "zahlungsinformationen", "skonto")),                                                                                       None,        12),
        ("zahlungsbedingungen", "Zahlungsbedingungen",    lambda d, e, raw: e.payment_terms if e else None,             None,        42),
        ("kommentar",           "Kommentar",              lambda d, e, raw: d.comment,                                  None,        30),
        ("importiert_am",       "Importiert am",          lambda d, e, raw: _dt(d.created_at),                          "datetime",  18),
    ]

    # Positionen-Sheet
    POSITION_DEFS = [
        ("beleg_nr",            "Beleg-Nr.",              lambda d, e, p, tm: d.id,                                         "center",  10),
        ("rechnungsnr",         "Rechnungsnr.",           lambda d, e, p, tm: e.invoice_number if e else None,              None,      22),
        ("lieferant",           "Lieferant",              lambda d, e, p, tm: e.vendor_id if e else None,                   None,      30),
        ("pos_nr",              "Pos.",                   lambda d, e, p, tm: p.position_index + 1,                         "center",   8),
        ("artikelbezeichnung",  "Artikelbezeichnung",     lambda d, e, p, tm: p.product_name or p.product_description,      None,      52),
        ("artikelnummer",       "Artikelnummer",          lambda d, e, p, tm: p.article_number,                             None,      20),
        ("menge",               "Menge",                  lambda d, e, p, tm: _f(p.quantity),                               None,      12),
        ("einheit",             "Einheit",                lambda d, e, p, tm: p.unit,                                       None,      12),
        ("einzelpreis",         "EP Netto (€)",           lambda d, e, p, tm: _f(p.unit_price_netto),                       "eur",     17),
        ("gesamtpreis",         "EP Brutto (€)",          lambda d, e, p, tm: _f(p.unit_price_brutto),                      "eur",     17),
        ("steuersatz",          "Steuersatz (%)",         lambda d, e, p, tm: tm.get(p.position_index),                     "pct",     14),
        ("nachlass",            "Nachlass",               lambda d, e, p, tm: p.discount,                                   None,      22),
    ]

    # Nur aktive Felder in der konfigurierten Reihenfolge
    inv_defs  = [d for d in INVOICE_DEFS  if d[0] in invoice_fields]
    pos_defs  = [d for d in POSITION_DEFS if d[0] in position_fields]

    # ── Sheet 1: Rechnungen ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Rechnungen"

    num_cols = max(len(inv_defs), 1)
    ws1.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    ws1["A1"] = f"Export: {batch.company_name} {batch.year}"
    ws1["A1"].font = Font(name="Arial", size=13, bold=True)
    ws1["A1"].alignment = Alignment(vertical="center")
    ws1.row_dimensions[1].height = 24

    _header_row(ws1, [d[1] for d in inv_defs], row=2)

    for r, doc in enumerate(docs, start=3):
        ex   = doc.extraction
        fill = _row_fill(r)
        raw  = _parse_raw(ex.raw_response if ex else None)

        for c, (key, header, val_fn, fmt, width) in enumerate(inv_defs, 1):
            val  = val_fn(doc, ex, raw)
            cell = ws1.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.fill = fill
            if fmt == "eur":          cell.number_format = fmt_eur
            elif fmt == "pct":        cell.number_format = fmt_pct
            elif fmt == "date":       cell.number_format = fmt_date
            elif fmt == "datetime":   cell.number_format = fmt_dt
            elif fmt == "center":     cell.alignment = center

    _set_widths(ws1, [d[4] for d in inv_defs])
    ws1.freeze_panes = "A3"

    # ── Sheet 2: Positionen ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Positionen")
    _header_row(ws2, [d[1] for d in pos_defs])

    r = 2
    for doc in docs:
        ex  = doc.extraction
        raw = _parse_raw(ex.raw_response if ex else None)
        raw_pos_list = raw.get("positionen") or []
        tax_map = {i: _f(p.get("steuersatz")) for i, p in enumerate(raw_pos_list)}

        for pos in doc.order_positions:
            fill = _row_fill(r)
            for c, (key, header, val_fn, fmt, width) in enumerate(pos_defs, 1):
                val  = val_fn(doc, ex, pos, tax_map)
                cell = ws2.cell(row=r, column=c, value=val)
                cell.font = data_font
                cell.fill = fill
                if fmt == "eur":      cell.number_format = fmt_eur
                elif fmt == "pct":    cell.number_format = fmt_pct
                elif fmt == "center": cell.alignment = center
            r += 1

    if r == 2:
        ws2.cell(row=2, column=1, value="Keine Positionen vorhanden").font = Font(
            name="Arial", size=10, color="888888", italic=True
        )

    _set_widths(ws2, [d[4] for d in pos_defs])
    ws2.freeze_panes = "A2"

    # ── Bytes zurückgeben ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
